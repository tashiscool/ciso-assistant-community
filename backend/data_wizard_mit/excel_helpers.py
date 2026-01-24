# MIT License - See LICENSE-MIT.txt in repository root
"""
Excel Helpers - Clean-room MIT implementation

Generic Excel reading/writing utilities for GRC data import/export.
"""

from io import BytesIO
from typing import List, Dict, Any, Optional, Iterator, Union
from dataclasses import dataclass, field


@dataclass
class CellValue:
    """Represents a cell value with optional formatting."""
    value: Any
    row: int
    column: int
    formula: Optional[str] = None
    hyperlink: Optional[str] = None


@dataclass
class SheetData:
    """Represents data from an Excel sheet."""
    name: str
    headers: List[str] = field(default_factory=list)
    rows: List[Dict[str, Any]] = field(default_factory=list)
    raw_rows: List[List[Any]] = field(default_factory=list)


class ExcelReader:
    """
    Generic Excel file reader with support for multiple formats.

    Provides utilities for reading Excel files and extracting data
    in a structured format suitable for GRC imports.
    """

    def __init__(self, file_content: Union[bytes, BytesIO]):
        """
        Initialize reader with file content.

        Args:
            file_content: Excel file as bytes or BytesIO
        """
        self._content = file_content if isinstance(file_content, BytesIO) else BytesIO(file_content)
        self._workbook = None
        self._sheet_names: List[str] = []

    def load(self) -> 'ExcelReader':
        """
        Load the workbook. Requires openpyxl to be installed.

        Returns:
            Self for method chaining
        """
        try:
            from openpyxl import load_workbook
            self._workbook = load_workbook(self._content, data_only=True)
            self._sheet_names = self._workbook.sheetnames
        except ImportError:
            raise ImportError("openpyxl is required for Excel operations")
        return self

    @property
    def sheet_names(self) -> List[str]:
        """Return list of sheet names in the workbook."""
        return self._sheet_names.copy()

    def get_sheet(self, name: str) -> Optional[Any]:
        """
        Get a sheet by exact name.

        Args:
            name: Exact sheet name

        Returns:
            Sheet object or None if not found
        """
        if self._workbook and name in self._workbook.sheetnames:
            return self._workbook[name]
        return None

    def get_sheet_data(self, name: str) -> Optional[SheetData]:
        """
        Get sheet data as structured object.

        Args:
            name: Sheet name

        Returns:
            SheetData object or None if sheet not found
        """
        sheet = self.get_sheet(name)
        if sheet is None:
            return None

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return SheetData(name=name)

        # First row as headers
        headers = [
            str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]

        # Data rows as dicts
        data_rows = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = cell
            data_rows.append(row_dict)

        return SheetData(
            name=name,
            headers=headers,
            rows=data_rows,
            raw_rows=[list(r) for r in rows]
        )

    def find_sheet_by_prefix(self, prefix: str) -> Optional[Any]:
        """
        Find a sheet whose name starts with prefix.

        Args:
            prefix: Prefix to match (e.g., "1.1" for "1.1 Study")

        Returns:
            First matching sheet or None
        """
        if self._workbook is None:
            return None
        for name in self._workbook.sheetnames:
            if name.startswith(prefix):
                return self._workbook[name]
        return None

    def find_sheets_by_prefix(self, prefix: str) -> List[tuple]:
        """
        Find all sheets matching a prefix.

        Args:
            prefix: Prefix to match

        Returns:
            List of (name, sheet) tuples
        """
        if self._workbook is None:
            return []
        return [
            (name, self._workbook[name])
            for name in self._workbook.sheetnames
            if name.startswith(prefix)
        ]

    def close(self):
        """Close the workbook."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None


class ExcelWriter:
    """
    Generic Excel file writer for GRC data exports.

    Provides utilities for creating Excel files with structured data,
    formatting, and multiple sheets.
    """

    def __init__(self):
        """Initialize the writer."""
        self._workbook = None
        self._sheets: Dict[str, Any] = {}

    def create(self) -> 'ExcelWriter':
        """
        Create a new workbook. Requires openpyxl.

        Returns:
            Self for method chaining
        """
        try:
            from openpyxl import Workbook
            self._workbook = Workbook()
            # Remove default sheet
            if "Sheet" in self._workbook.sheetnames:
                del self._workbook["Sheet"]
        except ImportError:
            raise ImportError("openpyxl is required for Excel operations")
        return self

    def add_sheet(self, name: str) -> Any:
        """
        Add a new sheet to the workbook.

        Args:
            name: Sheet name

        Returns:
            The created sheet object
        """
        if self._workbook is None:
            raise ValueError("Workbook not created. Call create() first.")
        sheet = self._workbook.create_sheet(name)
        self._sheets[name] = sheet
        return sheet

    def write_headers(self, sheet_name: str, headers: List[str], row: int = 1):
        """
        Write header row to a sheet.

        Args:
            sheet_name: Target sheet name
            headers: List of header strings
            row: Row number (1-indexed)
        """
        sheet = self._sheets.get(sheet_name)
        if sheet is None:
            sheet = self.add_sheet(sheet_name)
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=col, value=header)

    def write_row(self, sheet_name: str, values: List[Any], row: int):
        """
        Write a data row to a sheet.

        Args:
            sheet_name: Target sheet name
            values: List of cell values
            row: Row number (1-indexed)
        """
        sheet = self._sheets.get(sheet_name)
        if sheet is None:
            sheet = self.add_sheet(sheet_name)
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=value)

    def write_data(self, sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]):
        """
        Write headers and data rows to a sheet.

        Args:
            sheet_name: Target sheet name
            headers: List of header strings (also used as dict keys)
            rows: List of row dictionaries
        """
        self.write_headers(sheet_name, headers)
        for row_num, row_data in enumerate(rows, start=2):
            values = [row_data.get(h, "") for h in headers]
            self.write_row(sheet_name, values, row_num)

    def save(self, path_or_buffer: Union[str, BytesIO]) -> Optional[bytes]:
        """
        Save the workbook.

        Args:
            path_or_buffer: File path or BytesIO buffer

        Returns:
            Bytes if buffer provided, None if saved to file
        """
        if self._workbook is None:
            raise ValueError("Workbook not created")

        if isinstance(path_or_buffer, BytesIO):
            self._workbook.save(path_or_buffer)
            path_or_buffer.seek(0)
            return path_or_buffer.read()
        else:
            self._workbook.save(path_or_buffer)
            return None

    def close(self):
        """Close the workbook."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None


# Convenience functions

def find_sheet_by_name(workbook, names: List[str]) -> Optional[Any]:
    """
    Find a sheet by trying multiple possible names.

    Args:
        workbook: openpyxl Workbook object
        names: List of possible sheet names to try

    Returns:
        First matching sheet or None
    """
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def find_sheet_by_prefix(workbook, prefix: str) -> Optional[Any]:
    """
    Find a sheet whose name starts with prefix.

    Args:
        workbook: openpyxl Workbook object
        prefix: Prefix to match

    Returns:
        First matching sheet or None
    """
    for name in workbook.sheetnames:
        if name.startswith(prefix):
            return workbook[name]
    return None


def get_sheet_rows(sheet, header_row: int = 1) -> List[Dict[str, Any]]:
    """
    Extract rows from a sheet as list of dicts.

    Args:
        sheet: openpyxl sheet object
        header_row: Row number containing headers (1-indexed)

    Returns:
        List of row dictionaries with header keys
    """
    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < header_row + 1:
        return []

    # Get headers from specified row
    header_idx = header_row - 1
    headers = [
        str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
        for i, h in enumerate(rows[header_idx])
    ]

    # Extract data rows
    result = []
    for row in rows[header_row:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = cell
        result.append(row_dict)

    return result


def parse_multiline(value: Any) -> List[str]:
    """
    Parse a multiline cell value into a list of strings.

    Args:
        value: Cell value (string or other)

    Returns:
        List of non-empty lines
    """
    if not value:
        return []
    if isinstance(value, str):
        return [line.strip() for line in value.split("\n") if line.strip()]
    return [str(value)]


def normalize_header(header: str) -> str:
    """
    Normalize a header string for consistent key access.

    Args:
        header: Header string

    Returns:
        Normalized lowercase string with underscores
    """
    if not header:
        return ""
    return str(header).strip().lower().replace(" ", "_").replace("-", "_")


def coalesce(*values) -> Any:
    """
    Return first non-None value from arguments.

    Args:
        *values: Values to check

    Returns:
        First non-None value or None if all are None
    """
    for v in values:
        if v is not None:
            return v
    return None
