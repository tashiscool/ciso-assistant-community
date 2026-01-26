"""
RMF Export Service

Comprehensive export service for RMF Operations including:
- CKL file downloads (single and batch)
- Excel exports (checklists, findings, compliance)
- POAM exports
- Test plan exports
- Nessus summary exports
"""

import logging
import zipfile
from io import BytesIO
from typing import Dict, List, Any, Optional
from datetime import datetime
import uuid

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .ckl_exporter import CKLExporter

logger = logging.getLogger(__name__)


class RMFExportService:
    """
    Service for exporting RMF data in various formats.
    """

    def __init__(self):
        self.ckl_exporter = CKLExporter()

        # Standard styling
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.wrap_alignment = Alignment(vertical='top', wrap_text=True)

    # =========================================================================
    # CKL Exports
    # =========================================================================

    def export_checklist_ckl(self, checklist) -> HttpResponse:
        """
        Export a single checklist as CKL XML file.
        """
        ckl_data = self._build_ckl_data(checklist)
        ckl_content = self.ckl_exporter.export_to_ckl(ckl_data)

        filename = f"{checklist.hostName}_{checklist.stigType}.ckl"
        response = HttpResponse(ckl_content, content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_system_checklists_zip(self, system_group, checklists) -> HttpResponse:
        """
        Export all checklists for a system group as a ZIP file containing CKL files.
        """
        buffer = BytesIO()

        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for checklist in checklists:
                try:
                    ckl_data = self._build_ckl_data(checklist)
                    ckl_content = self.ckl_exporter.export_to_ckl(ckl_data)
                    filename = f"{checklist.hostName}_{checklist.stigType}.ckl"
                    zip_file.writestr(filename, ckl_content)
                except Exception as e:
                    logger.error(f"Error exporting checklist {checklist.id}: {e}")
                    continue

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{system_group.name}_checklists.zip"'
        return response

    def _build_ckl_data(self, checklist) -> Dict[str, Any]:
        """Build CKL data structure from checklist aggregate."""
        asset_info = checklist.assetInfo or {}

        return {
            'version': checklist.version or '1.0',
            'asset': {
                'host_name': checklist.hostName,
                'host_ip': asset_info.get('ip_addresses', [''])[0] if asset_info.get('ip_addresses') else '',
                'host_mac': asset_info.get('mac_addresses', [''])[0] if asset_info.get('mac_addresses') else '',
                'host_fqdn': asset_info.get('fqdn', ''),
                'tech_area': asset_info.get('tech_area', ''),
                'asset_type': checklist.asset_type or 'Computing',
                'role': asset_info.get('role', 'Member Server'),
                'web_or_database': checklist.isWebDatabase,
                'web_db_site': checklist.webDatabaseSite or '',
                'web_db_instance': checklist.webDatabaseInstance or ''
            },
            'stigs': {
                'stig_info': [
                    {'sid_name': 'version', 'sid_data': checklist.stigRelease or '1'},
                    {'sid_name': 'title', 'sid_data': checklist.stigType or ''},
                ],
                'stig_data': []
            },
            'vulnerabilities': self._build_vuln_list(checklist)
        }

    def _build_vuln_list(self, checklist) -> List[Dict[str, Any]]:
        """Build vulnerability list from checklist findings."""
        from ..repositories.vulnerability_finding_repository import VulnerabilityFindingRepository

        repo = VulnerabilityFindingRepository()
        findings = repo.find_by_checklist(checklist.id)

        vuln_list = []
        for finding in findings:
            status = finding.vulnerability_status or {}
            vuln_list.append({
                'status': self._map_status_to_ckl(status.get('status', 'not_reviewed')),
                'finding_details': status.get('finding_details', ''),
                'comments': status.get('comments', ''),
                'severity_override': status.get('severity_override', ''),
                'severity_justification': status.get('severity_justification', ''),
                'stig_data': [
                    {'attribute': 'Vuln_Num', 'data': finding.vulnId or ''},
                    {'attribute': 'Rule_ID', 'data': finding.ruleId or ''},
                    {'attribute': 'Severity', 'data': finding.severity.get('name', 'medium') if finding.severity else 'medium'},
                    {'attribute': 'Rule_Title', 'data': finding.ruleTitle or ''},
                ]
            })

        return vuln_list

    def _map_status_to_ckl(self, status: str) -> str:
        """Map internal status to CKL status format."""
        status_map = {
            'open': 'Open',
            'not_a_finding': 'NotAFinding',
            'not_applicable': 'Not_Applicable',
            'not_reviewed': 'Not_Reviewed'
        }
        return status_map.get(status, 'Not_Reviewed')

    # =========================================================================
    # Excel Exports
    # =========================================================================

    def export_checklist_xlsx(self, checklist) -> HttpResponse:
        """
        Export a single checklist to Excel format.
        """
        from ..repositories.vulnerability_finding_repository import VulnerabilityFindingRepository

        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"

        # Headers
        headers = [
            "Vuln ID", "STIG ID", "Rule ID", "Rule Title", "Severity",
            "Status", "Finding Details", "Comments", "CCI References"
        ]
        self._write_headers(ws, headers)

        # Data
        repo = VulnerabilityFindingRepository()
        findings = repo.find_by_checklist(checklist.id)

        for row_idx, finding in enumerate(findings, start=2):
            status = finding.vulnerability_status or {}
            severity = finding.severity or {}
            ws.cell(row=row_idx, column=1, value=finding.vulnId)
            ws.cell(row=row_idx, column=2, value=finding.stigId)
            ws.cell(row=row_idx, column=3, value=finding.ruleId)
            ws.cell(row=row_idx, column=4, value=finding.ruleTitle)
            ws.cell(row=row_idx, column=5, value=severity.get('name', ''))
            ws.cell(row=row_idx, column=6, value=status.get('status', ''))
            ws.cell(row=row_idx, column=7, value=status.get('finding_details', ''))
            ws.cell(row=row_idx, column=8, value=status.get('comments', ''))
            ws.cell(row=row_idx, column=9, value=', '.join(finding.cciIds or []))

            # Apply border and alignment
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = self.wrap_alignment

        # Auto-size columns
        self._auto_size_columns(ws)

        # Return response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{checklist.hostName}_{checklist.stigType}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_system_xlsx(self, system_group, checklists, scores) -> HttpResponse:
        """
        Export system group data to Excel format with multiple sheets.
        """
        wb = Workbook()

        # Summary sheet
        self._create_summary_sheet(wb, system_group, checklists, scores)

        # Checklists sheet
        self._create_checklists_sheet(wb, checklists)

        # Findings summary sheet
        self._create_findings_summary_sheet(wb, checklists)

        # Return response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{system_group.name}_report.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _create_summary_sheet(self, wb, system_group, checklists, scores):
        """Create summary sheet."""
        ws = wb.active
        ws.title = "Summary"

        # Title
        ws['A1'] = f"System Group: {system_group.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Description: {system_group.description or 'N/A'}"

        # Statistics
        ws['A5'] = "Statistics"
        ws['A5'].font = Font(bold=True, size=12)

        stats_data = [
            ("Total Checklists", len(checklists)),
            ("Total Open Cat I", sum(s.totalCat1Open for s in scores if s)),
            ("Total Open Cat II", sum(s.totalCat2Open for s in scores if s)),
            ("Total Open Cat III", sum(s.totalCat3Open for s in scores if s)),
        ]

        for idx, (label, value) in enumerate(stats_data, start=6):
            ws.cell(row=idx, column=1, value=label)
            ws.cell(row=idx, column=2, value=value)

    def _create_checklists_sheet(self, wb, checklists):
        """Create checklists listing sheet."""
        ws = wb.create_sheet("Checklists")

        headers = ["Host Name", "STIG Type", "Version", "Release", "Status", "Asset Type"]
        self._write_headers(ws, headers)

        for row_idx, cl in enumerate(checklists, start=2):
            ws.cell(row=row_idx, column=1, value=cl.hostName)
            ws.cell(row=row_idx, column=2, value=cl.stigType)
            ws.cell(row=row_idx, column=3, value=cl.version)
            ws.cell(row=row_idx, column=4, value=cl.stigRelease)
            ws.cell(row=row_idx, column=5, value=cl.lifecycle_state)
            ws.cell(row=row_idx, column=6, value=cl.asset_type)

        self._auto_size_columns(ws)

    def _create_findings_summary_sheet(self, wb, checklists):
        """Create findings summary sheet."""
        ws = wb.create_sheet("Findings Summary")

        headers = ["Host Name", "STIG Type", "Cat I Open", "Cat II Open", "Cat III Open", "Not Reviewed"]
        self._write_headers(ws, headers)

        from ..repositories.vulnerability_finding_repository import VulnerabilityFindingRepository
        repo = VulnerabilityFindingRepository()

        for row_idx, cl in enumerate(checklists, start=2):
            findings = repo.find_by_checklist(cl.id)

            cat1_open = sum(1 for f in findings if f.severity and f.severity.get('category') == 'cat1'
                          and f.vulnerability_status and f.vulnerability_status.get('status') == 'open')
            cat2_open = sum(1 for f in findings if f.severity and f.severity.get('category') == 'cat2'
                          and f.vulnerability_status and f.vulnerability_status.get('status') == 'open')
            cat3_open = sum(1 for f in findings if f.severity and f.severity.get('category') == 'cat3'
                          and f.vulnerability_status and f.vulnerability_status.get('status') == 'open')
            not_reviewed = sum(1 for f in findings if f.vulnerability_status
                              and f.vulnerability_status.get('status') == 'not_reviewed')

            ws.cell(row=row_idx, column=1, value=cl.hostName)
            ws.cell(row=row_idx, column=2, value=cl.stigType)
            ws.cell(row=row_idx, column=3, value=cat1_open)
            ws.cell(row=row_idx, column=4, value=cat2_open)
            ws.cell(row=row_idx, column=5, value=cat3_open)
            ws.cell(row=row_idx, column=6, value=not_reviewed)

        self._auto_size_columns(ws)

    # =========================================================================
    # POAM Export
    # =========================================================================

    def export_poam_xlsx(self, system_group, checklists) -> HttpResponse:
        """
        Export POAM (Plan of Action and Milestones) to Excel format.
        """
        from ..repositories.vulnerability_finding_repository import VulnerabilityFindingRepository

        wb = Workbook()
        ws = wb.active
        ws.title = "POA&M"

        headers = [
            "POA&M ID", "Weakness Name", "Weakness Description", "Detector Source",
            "Asset Identifier", "Point of Contact", "Resources Required",
            "Remediation Plan", "Detection Date", "Scheduled Completion",
            "Milestones", "Status", "Original Severity", "Comments"
        ]
        self._write_headers(ws, headers)

        repo = VulnerabilityFindingRepository()
        row_idx = 2
        poam_counter = 1

        for cl in checklists:
            findings = repo.find_by_checklist(cl.id)

            # Only include open findings in POAM
            open_findings = [f for f in findings if f.vulnerability_status
                           and f.vulnerability_status.get('status') == 'open']

            for finding in open_findings:
                status = finding.vulnerability_status or {}
                severity = finding.severity or {}

                poam_id = f"V-{datetime.now().year}-{poam_counter:04d}"

                ws.cell(row=row_idx, column=1, value=poam_id)
                ws.cell(row=row_idx, column=2, value=finding.ruleTitle or finding.vulnId)
                ws.cell(row=row_idx, column=3, value=finding.ruleDiscussion or '')
                ws.cell(row=row_idx, column=4, value="STIG Assessment")
                ws.cell(row=row_idx, column=5, value=cl.hostName)
                ws.cell(row=row_idx, column=6, value="")  # POC to be filled
                ws.cell(row=row_idx, column=7, value="")  # Resources to be filled
                ws.cell(row=row_idx, column=8, value=finding.fixText or '')
                ws.cell(row=row_idx, column=9, value=datetime.now().strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=10, value="")  # Completion date to be filled
                ws.cell(row=row_idx, column=11, value="")  # Milestones to be filled
                ws.cell(row=row_idx, column=12, value="Open")
                ws.cell(row=row_idx, column=13, value=severity.get('name', '').upper())
                ws.cell(row=row_idx, column=14, value=status.get('comments', ''))

                # Apply styling
                for col in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = self.border
                    cell.alignment = self.wrap_alignment

                row_idx += 1
                poam_counter += 1

        self._auto_size_columns(ws)
        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{system_group.name}_POAM.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # =========================================================================
    # Test Plan Export
    # =========================================================================

    def export_test_plan_xlsx(self, system_group, checklists) -> HttpResponse:
        """
        Export Security Test Plan to Excel format.
        """
        wb = Workbook()

        # Overview sheet
        ws = wb.active
        ws.title = "Test Plan Overview"
        ws['A1'] = f"Security Test Plan: {system_group.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total STIGs: {len(checklists)}"

        # Test Cases sheet
        tc_ws = wb.create_sheet("Test Cases")
        headers = [
            "Test Case ID", "Category", "STIG Reference", "Test Title",
            "Description", "Preconditions", "Test Steps", "Expected Results",
            "Status", "Tester", "Execution Date", "Evidence", "Notes"
        ]
        self._write_headers(tc_ws, headers)

        from ..repositories.vulnerability_finding_repository import VulnerabilityFindingRepository
        repo = VulnerabilityFindingRepository()

        row_idx = 2
        tc_counter = 1

        for cl in checklists:
            findings = repo.find_by_checklist(cl.id)

            for finding in findings:
                tc_id = f"TC-{tc_counter:05d}"
                severity = finding.severity or {}

                tc_ws.cell(row=row_idx, column=1, value=tc_id)
                tc_ws.cell(row=row_idx, column=2, value=f"STIG-{severity.get('category', 'cat2').upper()}")
                tc_ws.cell(row=row_idx, column=3, value=f"{cl.stigType} - {finding.vulnId}")
                tc_ws.cell(row=row_idx, column=4, value=finding.ruleTitle or '')
                tc_ws.cell(row=row_idx, column=5, value=finding.ruleDiscussion or '')
                tc_ws.cell(row=row_idx, column=6, value=f"System: {cl.hostName}")
                tc_ws.cell(row=row_idx, column=7, value=finding.checkContent or '')
                tc_ws.cell(row=row_idx, column=8, value="Verify compliance with STIG requirement")
                tc_ws.cell(row=row_idx, column=9, value="Not Run")
                tc_ws.cell(row=row_idx, column=10, value="")
                tc_ws.cell(row=row_idx, column=11, value="")
                tc_ws.cell(row=row_idx, column=12, value="")
                tc_ws.cell(row=row_idx, column=13, value="")

                row_idx += 1
                tc_counter += 1

        self._auto_size_columns(tc_ws)
        tc_ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{system_group.name}_TestPlan.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # =========================================================================
    # Nessus Exports
    # =========================================================================

    def export_nessus_summary_xlsx(self, system_group, scans) -> HttpResponse:
        """
        Export Nessus scan summary to Excel format.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Summary statistics
        ws['A1'] = f"Nessus Scan Summary: {system_group.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        total_critical = sum(s.critical_count for s in scans)
        total_high = sum(s.high_count for s in scans)
        total_medium = sum(s.medium_count for s in scans)
        total_low = sum(s.low_count for s in scans)
        total_info = sum(s.info_count for s in scans)

        ws['A5'] = "Severity Totals"
        ws['A5'].font = Font(bold=True)
        ws['A6'] = "Critical"
        ws['B6'] = total_critical
        ws['A7'] = "High"
        ws['B7'] = total_high
        ws['A8'] = "Medium"
        ws['B8'] = total_medium
        ws['A9'] = "Low"
        ws['B9'] = total_low
        ws['A10'] = "Info"
        ws['B10'] = total_info

        # Scans list
        scans_ws = wb.create_sheet("Scans")
        headers = ["Filename", "Scan Date", "Hosts", "Critical", "High", "Medium", "Low", "Status"]
        self._write_headers(scans_ws, headers)

        for row_idx, scan in enumerate(scans, start=2):
            scans_ws.cell(row=row_idx, column=1, value=scan.filename)
            scans_ws.cell(row=row_idx, column=2, value=str(scan.scan_date) if scan.scan_date else '')
            scans_ws.cell(row=row_idx, column=3, value=scan.total_hosts)
            scans_ws.cell(row=row_idx, column=4, value=scan.critical_count)
            scans_ws.cell(row=row_idx, column=5, value=scan.high_count)
            scans_ws.cell(row=row_idx, column=6, value=scan.medium_count)
            scans_ws.cell(row=row_idx, column=7, value=scan.low_count)
            scans_ws.cell(row=row_idx, column=8, value=scan.processing_status)

        self._auto_size_columns(scans_ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{system_group.name}_Nessus_Summary.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # =========================================================================
    # Helper Methods
    # =========================================================================

    def _write_headers(self, ws, headers: List[str]):
        """Write styled headers to worksheet."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _auto_size_columns(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-size columns based on content."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width
